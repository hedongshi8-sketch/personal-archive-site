from __future__ import annotations

import json
import re
import shutil
import subprocess
from datetime import date, datetime
from pathlib import Path
from typing import Any

from docx import Document
from docx.oxml.ns import qn
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
import pdfplumber


ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public"
ASSETS = PUBLIC / "portfolio-assets"
OUT = PUBLIC / "portfolio-previews"

MAX_EXCEL_ROWS = 160
MAX_EXCEL_COLS = 40
MAX_DOCUMENT_BLOCKS = 180
MAX_TABLE_ROWS = 80
MAX_TABLE_COLS = 16
MAX_PDF_BLOCKS = 180
MAX_PDF_VISUAL_PAGES = 24

CONTENT_TYPE_EXTENSIONS = {
    "image/jpeg": "jpg",
    "image/jpg": "jpg",
    "image/png": "png",
    "image/gif": "gif",
    "image/bmp": "bmp",
    "image/tiff": "tif",
    "image/webp": "webp",
}


def public_path(path: Path) -> str:
    return path.relative_to(PUBLIC).as_posix()


def slug(value: str) -> str:
    normalized = re.sub(r"\s+", "-", value.strip().lower())
    normalized = re.sub(r"[^\w\-\u4e00-\u9fff]+", "-", normalized)
    normalized = re.sub(r"-{2,}", "-", normalized).strip("-")
    return normalized or "sheet"


def format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.6g}"
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def trim_matrix(rows: list[list[str]]) -> list[list[str]]:
    while rows and all(not cell for cell in rows[-1]):
        rows.pop()

    if not rows:
        return []

    last_col = 0
    for row in rows:
        for index, cell in enumerate(row):
            if cell:
                last_col = max(last_col, index + 1)

    return [row[:last_col] for row in rows]


def read_text(path: Path) -> str:
    for encoding in ("utf-8", "utf-8-sig", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def write_json(relative_path: str, payload: dict[str, Any]) -> None:
    target = OUT / relative_path
    target.parent.mkdir(parents=True, exist_ok=True)
    payload["generatedAt"] = datetime.now().strftime("%Y-%m-%d")
    target.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"wrote {target.relative_to(ROOT)}")


def reset_media_dir(item_id: str) -> Path:
    media_dir = OUT / "media" / item_id
    if media_dir.exists():
        shutil.rmtree(media_dir)
    media_dir.mkdir(parents=True, exist_ok=True)
    return media_dir


def media_public_path(path: Path) -> str:
    return path.relative_to(PUBLIC).as_posix()


def extension_from_content_type(content_type: str | None, fallback: str = "png") -> str:
    if not content_type:
        return fallback
    return CONTENT_TYPE_EXTENSIONS.get(content_type.lower(), fallback)


def get_anchor_position(anchor: Any) -> dict[str, int | None]:
    start = getattr(anchor, "_from", None)
    end = getattr(anchor, "to", None)
    return {
        "row": getattr(start, "row", None) + 1 if getattr(start, "row", None) is not None else None,
        "column": getattr(start, "col", None) + 1 if getattr(start, "col", None) is not None else None,
        "toRow": getattr(end, "row", None) + 1 if getattr(end, "row", None) is not None else None,
        "toColumn": getattr(end, "col", None) + 1 if getattr(end, "col", None) is not None else None,
    }


def find_pdftoppm() -> str | None:
    candidates = [
        ROOT / ".codex" / "pdftoppm.exe",
        Path.home() / ".cache" / "codex-runtimes" / "codex-primary-runtime" / "dependencies" / "native" / "poppler" / "Library" / "bin" / "pdftoppm.exe",
        Path.home() / ".cache" / "codex-runtimes" / "codex-primary-runtime" / "dependencies" / "bin" / "pdftoppm.cmd",
    ]
    for candidate in candidates:
        if candidate.exists():
            return str(candidate)
    return shutil.which("pdftoppm")


def build_pdf_page_images(item_id: str, source: Path, page_count: int) -> list[dict[str, Any]]:
    pdftoppm = find_pdftoppm()
    if not pdftoppm:
        return []

    media_dir = reset_media_dir(item_id)
    prefix = media_dir / "page"
    last_page = min(page_count, MAX_PDF_VISUAL_PAGES)

    subprocess.run(
        [
            pdftoppm,
            "-png",
            "-r",
            "128",
            "-f",
            "1",
            "-l",
            str(last_page),
            str(source),
            str(prefix),
        ],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )

    page_images: list[dict[str, Any]] = []
    for page_index, image_path in enumerate(sorted(media_dir.glob("page-*.png")), start=1):
        page_images.append(
            {
                "page": page_index,
                "src": media_public_path(image_path),
                "alt": f"第 {page_index} 页预览",
            }
        )

    return page_images


def extract_docx_paragraph_images(item_id: str, document: Document, paragraph: Paragraph, media_dir: Path, image_counter: int) -> tuple[list[dict[str, Any]], int]:
    blocks: list[dict[str, Any]] = []
    for blip in paragraph._p.xpath(".//a:blip"):
        relationship_id = blip.get(qn("r:embed")) or blip.get(qn("r:link"))
        if not relationship_id or relationship_id not in document.part.related_parts:
            continue

        image_part = document.part.related_parts[relationship_id]
        extension = extension_from_content_type(getattr(image_part, "content_type", None))
        image_counter += 1
        image_path = media_dir / f"doc-image-{image_counter}.{extension}"
        image_path.write_bytes(image_part.blob)
        blocks.append(
            {
                "type": "image",
                "src": media_public_path(image_path),
                "alt": f"文档图片 {image_counter}",
            }
        )

    return blocks, image_counter


def build_excel_preview(item_id: str, title: str, sources: list[Path]) -> None:
    sheets: list[dict[str, Any]] = []
    media_dir = reset_media_dir(item_id)
    image_counter = 0

    for source in sources:
        workbook = load_workbook(source, read_only=False, data_only=False)
        try:
            for worksheet in workbook.worksheets:
                max_row = min(worksheet.max_row or 0, MAX_EXCEL_ROWS)
                max_col = min(worksheet.max_column or 0, MAX_EXCEL_COLS)
                sheet_images: list[dict[str, Any]] = []
                sheet_id = slug(f"{source.stem}-{worksheet.title}")

                for image in getattr(worksheet, "_images", []):
                    try:
                        image_counter += 1
                        image_format = (getattr(image, "format", None) or "png").lower()
                        extension = "jpg" if image_format == "jpeg" else image_format
                        if extension not in {"png", "jpg", "jpeg", "gif", "bmp", "webp"}:
                            extension = "png"
                        image_path = media_dir / f"{sheet_id}-image-{image_counter}.{extension}"
                        image_path.write_bytes(image._data())
                        sheet_images.append(
                            {
                                "id": f"{sheet_id}-image-{image_counter}",
                                "src": media_public_path(image_path),
                                "alt": f"{worksheet.title} 图片 {len(sheet_images) + 1}",
                                "anchor": get_anchor_position(getattr(image, "anchor", None)),
                                "width": getattr(image, "width", None),
                                "height": getattr(image, "height", None),
                            }
                        )
                    except Exception as error:
                        print(f"warning: failed to extract image from {source.name}/{worksheet.title}: {error}")

                if max_row <= 0 or max_col <= 0:
                    if not sheet_images:
                        continue
                    rows = []
                else:
                    rows = [
                        [format_value(value) for value in row]
                        for row in worksheet.iter_rows(
                            min_row=1,
                            max_row=max_row,
                            min_col=1,
                            max_col=max_col,
                            values_only=True,
                        )
                    ]
                    rows = trim_matrix(rows)

                if not rows and not sheet_images:
                    continue

                sheets.append(
                    {
                        "id": sheet_id,
                        "workbookName": source.name,
                        "sheetName": worksheet.title,
                        "rowCount": worksheet.max_row,
                        "columnCount": worksheet.max_column,
                        "truncatedRows": (worksheet.max_row or 0) > MAX_EXCEL_ROWS,
                        "truncatedColumns": (worksheet.max_column or 0) > MAX_EXCEL_COLS,
                        "cells": rows,
                        "images": sheet_images,
                    }
                )
        finally:
            workbook.close()

    write_json(
        f"{item_id}.json",
        {
            "kind": "excel",
            "title": title,
            "sourceFiles": [public_path(source) for source in sources],
            "sheetCount": len(sheets),
            "rowLimit": MAX_EXCEL_ROWS,
            "columnLimit": MAX_EXCEL_COLS,
            "sheets": sheets,
        },
    )


def iter_docx_blocks(document: Document):
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, document)
        elif isinstance(child, CT_Tbl):
            yield Table(child, document)


def paragraph_block(paragraph: Paragraph) -> dict[str, Any] | None:
    text = paragraph.text.strip()
    if not text:
        return None

    style_name = paragraph.style.name if paragraph.style is not None else ""
    if style_name.startswith("Heading"):
        match = re.search(r"(\d+)", style_name)
        return {
            "type": "heading",
            "level": int(match.group(1)) if match else 2,
            "text": text,
        }

    if style_name in {"Title", "Subtitle"}:
        return {"type": "heading", "level": 1 if style_name == "Title" else 2, "text": text}

    return {"type": "paragraph", "text": text}


def paragraph_blocks(item_id: str, document: Document, paragraph: Paragraph, media_dir: Path, image_counter: int) -> tuple[list[dict[str, Any]], int]:
    blocks: list[dict[str, Any]] = []
    text_block = paragraph_block(paragraph)
    if text_block:
        blocks.append(text_block)

    image_blocks, image_counter = extract_docx_paragraph_images(item_id, document, paragraph, media_dir, image_counter)
    blocks.extend(image_blocks)
    return blocks, image_counter


def table_block(table: Table) -> dict[str, Any] | None:
    rows: list[list[str]] = []
    for row in table.rows[:MAX_TABLE_ROWS]:
        cells = [cell.text.strip().replace("\r", "\n") for cell in row.cells[:MAX_TABLE_COLS]]
        rows.append(cells)

    rows = trim_matrix(rows)
    if not rows:
        return None

    return {
        "type": "table",
        "rowCount": len(table.rows),
        "columnCount": len(table.columns),
        "truncatedRows": len(table.rows) > MAX_TABLE_ROWS,
        "truncatedColumns": len(table.columns) > MAX_TABLE_COLS,
        "rows": rows,
    }


def build_docx_preview(item_id: str, title: str, source: Path) -> None:
    document = Document(source)
    blocks: list[dict[str, Any]] = []
    media_dir = reset_media_dir(item_id)
    image_counter = 0

    for block in iter_docx_blocks(document):
        if isinstance(block, Table):
            next_block = table_block(block)
            if next_block:
                blocks.append(next_block)
        else:
            next_blocks, image_counter = paragraph_blocks(item_id, document, block, media_dir, image_counter)
            blocks.extend(next_blocks)
        if len(blocks) >= MAX_DOCUMENT_BLOCKS:
            break

    write_json(
        f"{item_id}.json",
        {
            "kind": "document",
            "title": title,
            "sourceFile": public_path(source),
            "blockLimit": MAX_DOCUMENT_BLOCKS,
            "truncatedBlocks": len(blocks) >= MAX_DOCUMENT_BLOCKS,
            "blocks": blocks,
        },
    )


def build_pdf_preview(item_id: str, title: str, source: Path) -> None:
    blocks: list[dict[str, Any]] = []

    with pdfplumber.open(source) as pdf:
        page_count = len(pdf.pages)
        page_images = build_pdf_page_images(item_id, source, page_count)
        for page_index, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=False) or ""
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            if not lines:
                continue

            blocks.append(
                {
                    "type": "heading",
                    "level": 2,
                    "text": f"第 {page_index} 页 / 共 {page_count} 页",
                }
            )

            for line in lines:
                if len(blocks) >= MAX_PDF_BLOCKS:
                    break

                if len(line) <= 34 and (
                    line.startswith(("一、", "二、", "三、", "四、", "五、", "六、"))
                    or line.endswith(("说明", "拆解", "目录", "内容摘要", "功能说明"))
                ):
                    blocks.append({"type": "heading", "level": 3, "text": line})
                else:
                    blocks.append({"type": "paragraph", "text": line})

            if len(blocks) >= MAX_PDF_BLOCKS:
                break

    write_json(
        f"{item_id}.json",
        {
            "kind": "document",
            "title": title,
            "sourceFile": public_path(source),
            "blockLimit": MAX_PDF_BLOCKS,
            "truncatedBlocks": len(blocks) >= MAX_PDF_BLOCKS,
            "pageImages": page_images,
            "truncatedPageImages": page_count > MAX_PDF_VISUAL_PAGES,
            "blocks": blocks,
        },
    )


def build_text_preview(item_id: str, title: str, source: Path, kind: str) -> None:
    write_json(
        f"{item_id}.json",
        {
            "kind": kind,
            "title": title,
            "sourceFile": public_path(source),
            "content": read_text(source),
        },
    )


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)

    build_excel_preview(
        "barbarq-main-sheet",
        "菇霸争夺战配置表",
        [ASSETS / "barbarq" / "sheets" / "野蛮人大作战2-菇霸争夺战.xlsx"],
    )
    build_excel_preview(
        "barbarq-related-sheet",
        "菇霸争夺战相关表格",
        [ASSETS / "barbarq" / "sheets" / "野蛮人大作战2-菇霸争夺战相关表格.xlsx"],
    )
    build_excel_preview(
        "barbarq-art-sheet",
        "菇霸争夺战美术需求表",
        [ASSETS / "barbarq" / "sheets" / "野蛮人大作战2-菇霸争夺战部分美术需求.xlsx"],
    )
    build_excel_preview(
        "system-planner-war-sheet",
        "战意 / 骑砍2 / 全面战争系统拆解案",
        [ASSETS / "system-planner" / "sheets" / "系统策划拆解案_战意_骑砍2_全面战争.xlsx"],
    )
    build_excel_preview(
        "game-town-config-sheets",
        "游戏小镇系统配置表合集",
        sorted((ASSETS / "game-town" / "sheets").glob("*.xlsx")),
    )
    build_pdf_preview(
        "barbarq-main-design",
        "野蛮人大作战2 - 菇霸争夺战策划案",
        ASSETS / "barbarq" / "docs" / "野蛮人大作战2-菇霸争夺战.pdf",
    )
    build_pdf_preview(
        "barbarq-art-requirement",
        "菇霸争夺战部分美术需求",
        ASSETS / "barbarq" / "docs" / "野蛮人大作战2-菇霸争夺战部分美术需求.pdf",
    )
    build_pdf_preview(
        "system-planner-portfolio",
        "系统策划实习生作品集",
        ASSETS / "system-planner" / "docs" / "01_作品集_系统策划实习生_最终投递版.pdf",
    )

    build_docx_preview(
        "game-town-design-doc",
        "游戏小镇方案完善版",
        ASSETS / "game-town" / "docs" / "游戏小镇方案V_0.2完善版(1).docx",
    )
    build_text_preview(
        "game-town-prototype-readme",
        "游戏小镇原型说明",
        ASSETS / "game-town" / "docs" / "README-原型说明.md",
        "markdown",
    )
    build_text_preview(
        "game-town-expanded-design",
        "游戏小镇方案补全文档",
        ASSETS / "game-town" / "docs" / "游戏小镇方案补全文档.md",
        "markdown",
    )

if __name__ == "__main__":
    main()
